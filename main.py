from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.label import Label
from kivy.core.window import Window
from kivy.config import Config


from fpdf import FPDF
from fpdf.fonts import FontFace
import re
import openpyxl
from datetime import date
import os


# сумма цифер числа до однозначного значения (принимает str, возвращает int)
def spec_sum(num):
    sum = 0
    if int(num) < 1:
        num = str(9 + int(num))
    while len(num) > 1:
        for x in num:
            sum += int(x)
            if sum >= 10:
                sum = int(str(sum)[0]) + int(str(sum)[1])
        num = str(sum)
    return int(num)


# расчет цифер
def calc_nums(input_date):
    sdate = input_date.split('.')

    day = spec_sum(sdate[0])
    month = spec_sum(sdate[1])
    year = spec_sum(sdate[2][2:]) if int(sdate[2]) < 2000 else int(spec_sum(sdate[2][2:])) + 1
    act = spec_sum(input_date.replace('.', ''))
    temp = abs(abs(day - month) - abs(day - spec_sum(sdate[2])))
    pp = temp if temp != 0 else 9
    num_data = {'day': day, 'month': month, 'year': year, 'act': act,
                'psy': spec_sum(str(year - 2)),
                'fam': spec_sum(str(act + 2)),
                'm_out': spec_sum(str(month - 2)),
                'comm': spec_sum(str(day + 2)),
                'abi': spec_sum(str(act + 1)),
                'happy': spec_sum(str(year + 1)),
                'goal_ts': spec_sum(str(day - 1)),
                'love': spec_sum(str(month - 1)),
                'm_in': spec_sum(str(day + 1)),
                'care': spec_sum(str(month - 3)),
                'poss': spec_sum(str(act + 3)),
                'wv': spec_sum(str(year - 1)),
                'vector': spec_sum(str(act * 2 -1)),
                'key_goal': spec_sum(str(day + month)),
                'pp': pp}
    return num_data

#
def calc_competence_matrix(input_date):
    matrix = [[' ', ' ', ' '],
              [' ', ' ', ' '],
              [' ', ' ', ' ']]
    raw_matrix_lines = {'147': ['3', ''],
                    '258': ['4', ''],
                    '369': ['5', ''],
                    '123': ['6', ''],
                    '456': ['7', ''],
                    '789': ['8', ''],
                    '159': ['9', ''],
                    '357': ['10', ''],
                    '24': ['11', '1'],
                    '26': ['12', '1'],
                    '48': ['13', '1'],
                    '68': ['14', '1'],
                    '89': ['15', '1']}
    n = 0
    for x in range(3):
        for y in range(2, -1, -1):
            n += 1
            dc = input_date.count(str(n))
            if dc != 0:
                matrix[y][x] = str(n) * dc
                for ln in raw_matrix_lines.keys():
                    if ln.find(str(n)) != -1:
                        raw_matrix_lines[ln][1] += '1'

    filtered_matrix_lines = dict(filter(lambda ln: ln[1][1] == '111', raw_matrix_lines.items()))
    return matrix, filtered_matrix_lines

# вывод текста
def pdf_print_text(pdf, text, padding=True, **kwargs):
    if padding:
        pdf.multi_cell(w=0, h=pdf.font_size * 1.59, text=text, new_x='LMARGIN', new_y='NEXT', padding=1.5, **kwargs)
    else:
        pdf.multi_cell(w=0, h=pdf.font_size * 1.59, text=text, new_x='LMARGIN', new_y='NEXT', **kwargs)


def pdf_print_colon_text(pdf, text1, text2):
    pdf.set_text_color(57, 100, 151)
    pdf.set_font(style='BI')
    pdf.write(h=pdf.font_size * 1.59, text=text1)
    pdf.set_font(style='')
    pdf.set_text_color(0, 0, 0)
    pdf.write(h=pdf.font_size * 1.59, text=text2)
    pdf.set_font(style='')
    pdf.set_xy(x=pdf.l_margin, y=pdf.y + 10)

def pdf_print_colored_text(pdf, strokes, end_padding=False):

    for stroke in strokes:
        text = stroke[0]
        color = stroke[1]
        style = stroke[2]

        pdf.set_text_color(color)
        pdf.set_font(style=style)
        pdf.write(h=pdf.font_size * 1.59, text=text)

    pdf.set_text_color(0, 0, 0)
    pdf.set_font(style='')
    if end_padding:
        pdf.set_xy(x=pdf.l_margin, y=pdf.y + 10)

# вывод блока (заголовок, подзаголовок, текст(опционально))
def pdf_print_block(pdf, title, subtitle=None, text=None):
    pdf.cell(0, 10, '', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font(style='BI', size=12)
    pdf.set_text_color(0, 49, 83)
    pdf.multi_cell(w=0, h=pdf.font_size * 1.59, text=title, new_x='LMARGIN', new_y='NEXT', padding=1.5, align='C')
    if subtitle is not None:
        pdf.set_font(style='I', size=11)
        pdf.set_text_color(122, 122, 122)
        pdf.multi_cell(w=0, h=pdf.font_size * 1.59, text=subtitle, new_x='LMARGIN', new_y='NEXT', padding=1.5, align='C')
    pdf.set_text_color(0, 0, 0)
    pdf.set_font(style='', size=11)
    if text is not None:
        pdf.multi_cell(w=0, h=pdf.font_size * 1.59, text=text, new_x='LMARGIN', new_y='NEXT', padding=1.5)
    pdf.cell(0, 4, '', new_x='LMARGIN', new_y='NEXT')

# вывод таблицы со списком (в excel все элементы списка в одной ячейке, каждый с новой строки)
def pdf_print_table_list(pdf, title1, title2, table_text1, table_text2):
    pdf.cell(0, 3, '', new_x='LMARGIN', new_y='NEXT')
    headings_style = FontFace(color=(255, 255, 255), fill_color=(0, 49, 83))
    with pdf.table(line_height=pdf.font_size * 2, padding=5, headings_style=headings_style) as table:
        row = table.row()
        pdf.set_font(style='BI')
        row.cell(text=title1, align='C')
        row.cell(text=title2, align='C')
        pdf.set_font(style='')

        row = table.row()

        row.cell(text=make_list(table_text1), align='L')
        row.cell(text=make_list(table_text2), align='L')
    pdf.cell(0, 5, '', new_x='LMARGIN', new_y='NEXT')

def make_list(text):
    return text.replace('\n', '').replace('*', '•  ', 1).replace(' *', '\n•  ')


def load_DB():
    return openpyxl.load_workbook(filename='PsyCalcData.xlsx')


# проверка введенной даты
def check_date(input_date):
    regexp = r'\d{2}\.\d{2}\.\d{4}'
    if re.match(regexp, input_date) and len(input_date) == 10:
        return True
    else:
        return False


def DB_get_value(sheet, cell_x, cell_y):
    value = sheet[cell_x + str(cell_y + 2)].value
    if value is not None:
        return value
    else:
        return "Не заполнено"


def create_pdf(input_date, name, DB):
    # Задание настроек файла
    pdf = FPDF('P', 'mm', 'A4')
    pdf.add_page()
    pdf.add_font('DejaVu', '', 'font/DejaVuSerif.ttf')
    pdf.add_font('DejaVu', 'B', 'font/DejaVuSerif-Bold.ttf')
    pdf.add_font('DejaVu', 'I', 'font/DejaVuSerif-Italic.ttf')
    pdf.add_font('DejaVu', 'BI', 'font/DejaVuSerif-BoldItalic.ttf')
    pdf.set_font('DejaVu', '', size=11)
    pdf.set_margins(15, 10, 15)
    pdf.set_auto_page_break(auto=True, margin=15)

    blue = (57, 100, 151)
    # blue = (0, 128, 0)
    # blue = (55, 69, 43)
    red = (192, 0, 0)
    # red = (37, 57, 6)
    # red = (1, 51, 0)
    grey = (122, 122, 122)
    dark_blue = (0, 49, 83)
    black = (0, 0, 0)

    # Расчет основных параметров
    num_data = calc_nums(input_date)

    # Вывод данных
    pdf.set_font(style='B', size=15)
    pdf_print_text(pdf, 'Цифровая психология', align='C')
    pdf.set_font(style='BI', size=12)
    pdf.set_text_color(red)
    pdf_print_text(pdf, f'{name}, дата рождения {input_date}, {num_data["day"]} / {num_data["act"]}', align='C')
    pdf.set_text_color(dark_blue)
    pdf_print_text(pdf, f'Архетип личности - {num_data["day"]}', align='C')
    pdf.set_font(style='I', size=11)
    pdf.set_text_color(grey)
    pdf_print_text(pdf, '(психическая предрасположенность, определяющая поведение и мышление. Психические и поведенческие программы личности)', align='C')
    pdf.cell(0, 4, '', new_x='LMARGIN', new_y='NEXT')
    pdf.set_text_color(blue)
    pdf.set_font(style='BI')
    pdf_print_text(pdf, f'Преобладающий тип мышления: {DB_get_value(DB["Архетип личности"], "B", num_data["day"])}', align='C')
    pdf.set_text_color(0, 0, 0)
    pdf.set_font(style='')
    pdf_print_text(pdf, DB_get_value(DB['Архетип личности'], "C", num_data["day"]))
    pdf_print_table_list(pdf,
                         'В позитивном аспекте (+)',
                         'В негативном аспекте (-)',
                         DB_get_value(DB["Архетип личности"], "D", num_data["day"]),
                         DB_get_value(DB["Архетип личности"], "E", num_data["day"]))
    ##
    pdf.set_font(style='BI')
    pdf.set_text_color(blue)
    pdf.write(h=pdf.font_size * 1.59, text='Вектор эго направлен ')
    pdf.set_text_color(red)
    pdf.write(h=pdf.font_size * 1.59, text=DB_get_value(DB["Архетип личности"], "F", num_data["day"]))
    pdf.set_xy(x=pdf.l_margin, y=pdf.y + 10)
    ##
    pdf.set_text_color(blue)
    pdf.write(h=pdf.font_size * 1.59, text='Чего хотите: ')
    pdf.set_text_color(red)
    pdf.write(h=pdf.font_size * 1.59, text=DB_get_value(DB["Архетип личности"], "G", num_data["day"]))
    pdf.set_xy(x=pdf.l_margin, y=pdf.y + 10)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font(style='')
    ##
    pdf_print_table_list(pdf,
                         'Эго наслаждается',
                         'Причины разрушения',
                         DB_get_value(DB["Архетип личности"], "H", num_data["day"]),
                         DB_get_value(DB["Архетип личности"], "I", num_data["day"]))
    ##
    pdf_print_colored_text(pdf, [['Триггеры ', blue, 'BI'], ['(ситуации, вызывающие негативные эмоции):', grey, 'I']], True)
    pdf_print_text(pdf, make_list(DB_get_value(DB["Архетип личности"], "J", num_data["day"])))
    pdf.cell(0, 3, '', new_x='LMARGIN', new_y='NEXT')
    pdf_print_colon_text(pdf, 'Психосоматика болезней: ', DB_get_value(DB['Архетип личности'], 'K', num_data["day"]))
    ##
    pdf_print_block(pdf, f'Действия/врожденная форма поведения – {num_data["act"]}', '(неосознаваемые поведенческие паттерны. Готовность действовать определённым образом)')
    ##
    pdf_print_table_list(pdf, 'Конструктивные действия (форма поведения)',
                         'Деструктивные действия (форма поведения)',
                         DB_get_value(DB["Действия"], 'B', num_data['act']),
                         DB_get_value(DB["Действия"], 'C', num_data['act']))
    ##
    pdf_print_block(pdf, 'Матрица психических процессов личности', '(психические процессы личности, выступающие регуляторами поведения или восприятия в основных сферах жизни)')
    with pdf.table(line_height=1.59 * pdf.font_size, first_row_as_headings=False, text_align="C") as table:
        row = table.row()
        row.cell('Месяц\n' + str(num_data['month']))
        row.cell('День \n' + str(num_data['day']))
        row.cell('Год \n' + str(num_data['year']))
        row.cell('Действия \n' + str(num_data['act']))

        row = table.row()
        row.cell('Психика \n' + str(num_data['psy']))
        row.cell('Семья/близкие \n' + str(num_data['fam']))
        row.cell('Деньги/расходы \n' + str(num_data['m_out']))
        row.cell('Коммуникация \n' + str(num_data['comm']))

        row = table.row()
        row.cell('Способности \n' + str(num_data['abi']))
        row.cell('Счастье/комфорт \n' + str(num_data['happy']))
        row.cell('Задача для ТС \n' + str(num_data['goal_ts']))
        row.cell('Любовь \n' + str(num_data['love']))

        row = table.row()
        row.cell('Деньги/доходы \n' + str(num_data['m_in']))
        row.cell('Забота \n' + str(num_data['care']))
        row.cell('Возможности \n' + str(num_data['poss']))
        row.cell('Мировоззрение \n' + str(num_data['wv']))
    ##
    pdf_print_block(pdf, f'Психическая энергия - {num_data["psy"]}',
                    '(Востребованность в социуме всецело зависит от того под влиянием какой энергии находится человек, а также от её уровня)')
    pdf_print_colon_text(pdf, 'Образ в социуме: ', DB_get_value(DB['Психическая энергия'], 'B', num_data['psy']))
    pdf_print_colon_text(pdf, 'Внешний вид в социуме: ', DB_get_value(DB['Психическая энергия'], 'C', num_data['psy']))
    ##
    pdf.cell(0, 3, '', new_x='LMARGIN', new_y='NEXT')
    headings_style = FontFace(color=(255, 255, 255), fill_color=(0, 49, 83), emphasis='BI')
    with pdf.table(line_height=1.59 * pdf.font_size, headings_style=headings_style, borders_layout='NO_HORIZONTAL_LINES') as table:
        row = table.row()
        row.cell('Низкий уровень психической энергии', align="C", padding=5)
        row.cell('Высокий уровень психической энергии', align="C", padding=5)

        row = table.row()
        pdf.set_text_color(blue)
        row.cell('Низкая востребованность в социуме:', align="L", v_align='T', padding=(5, 5, 0, 5))
        row.cell('Высокая востребованность в социуме:', align="L", v_align='T', padding=(5, 5, 0, 5))
        pdf.set_font(style='')
        pdf.set_text_color(0, 0, 0)

        row = table.row()
        row.cell(DB_get_value(DB["Психическая энергия"], "D", num_data["psy"]), align="L", v_align='T', padding=(3, 5, 0, 5))
        row.cell(DB_get_value(DB["Психическая энергия"], "F", num_data["psy"]), align="L", v_align='T', padding=(3, 5, 0, 5))

        row = table.row()
        pdf.set_font(style='I')
        pdf.set_text_color(blue)
        row.cell('Внутренние ощущения:', align="L", v_align='T', padding=(5, 5, 0, 5))
        row.cell('Внутренние ощущения:', align="L", v_align='T', padding=(5, 5, 0, 5))
        pdf.set_font(style='')
        pdf.set_text_color(0, 0, 0)

        row = table.row()
        row.cell(DB_get_value(DB["Психическая энергия"], "E", num_data["psy"]), align="L", v_align='T', padding=(3, 5, 5, 5))
        row.cell(DB_get_value(DB["Психическая энергия"], "G", num_data["psy"]), align="L", v_align='T', padding=(3, 5, 5, 5))

        row = table.row()
        pdf.set_font(style='BI')
        row.cell(DB_get_value(DB["Психическая энергия"], "H", num_data["psy"]), colspan=2, style=headings_style, align="C", padding=5)
        pdf.set_font(style='')
    pdf.cell(0, 3, '', new_x='LMARGIN', new_y='NEXT')
    ##
    pdf_print_block(pdf, f'Семья/близкие люди - {num_data["fam"]}', 'близкое окружение: родственники, родители, дети, супруг(а)')
    with pdf.table(line_height=1.59 * pdf.font_size, headings_style=headings_style, padding=5) as table:
        row = table.row()
        row.cell('Характер взаимоотношений с близкими', align='C')
        row.cell('Чего ждете/хотите от близких людей', align='C')

        row = table.row()
        row.cell(DB_get_value(DB['Семья'], 'B', num_data['fam']), align='L')
        row.cell(DB_get_value(DB['Семья'], 'C', num_data['fam']), align='L')

        row = table.row()
        row.cell(DB_get_value(DB['Семья'], 'D', num_data['fam']), align='C', colspan=2, style=headings_style)
    ##
    pdf_print_block(pdf, f'Деньги/расходы - {num_data["m_out"]}', '(на что предпочитаете тратить деньги)', DB_get_value(DB['Деньги расходы'], 'B', num_data['m_out']))
    ##
    pdf_print_block(pdf, f'Коммуникация - {num_data["comm"]}', '(принцип взаимодействия в социуме/принцип деления людей)', DB_get_value(DB['Коммуникация'], 'B', num_data['comm']))
    ##
    pdf_print_block(pdf, f'Способности - {num_data["abi"]}', '(Ваши способности и/или интуиция)')
    pdf_print_colon_text(pdf, 'Способности: ', DB_get_value(DB['Способности'], 'B', num_data['abi']))
    pdf_print_colored_text(pdf, [['Интуиция: ', red, 'BI'],
                                  [DB_get_value(DB['Способности'], 'C', num_data["abi"]), black, '']], True)
    ##
    pdf_print_block(pdf, f'Ощущение счастья - {num_data["happy"]}', '(зона психологического комфорта)')
    pdf_print_colon_text(pdf, 'Зона психологического комфорта: ', DB_get_value(DB['Ощущение счастья'], 'B', num_data['happy']))
    ##
    pdf_print_block(pdf, f'Задача для трансформации сознания - {num_data["goal_ts"]}', '(действия, проживая которые происходит трансформация сознания)')
    pdf_print_colored_text(pdf, [[DB_get_value(DB['Задача для ТС'], 'B', num_data['goal_ts']), red, 'BI']], True)
    pdf_print_text(pdf, DB_get_value(DB['Задача для ТС'], 'C', num_data['goal_ts']))
    pdf.cell(0, 3, '', new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(pdf.l_margin + 38)
    pdf_print_colored_text(pdf, [['Формула задачи (стратегия счастливой жизни)', blue, 'I']], end_padding=True)
    with pdf.table(line_height=1.59 * pdf.font_size, first_row_as_headings=False, text_align='C',
                   col_widths=(15, 2, 17, 2, 15, 2, 14)) as table:
        col1 = make_list(DB_get_value(DB['Задача для ТС'], 'D', num_data['goal_ts']))
        col2 = make_list(DB_get_value(DB['Задача для ТС'], 'E', num_data['goal_ts']))
        col3 = DB_get_value(DB['Задача для ТС'], 'F', num_data['goal_ts'])
        col4 = DB_get_value(DB['Задача для ТС'], 'G', num_data['goal_ts'])

        row = table.row()
        row.cell(col1, align='L')
        row.cell('+')
        row.cell(col2, align='L')
        row.cell('=')
        row.cell(col3)
        row.cell('⇒')
        row.cell(col4)
    pdf.cell(0, 6, '', new_x='LMARGIN', new_y='NEXT')
    pdf_print_colored_text(pdf, [['Аффирмация ', blue, 'I'],
                           ['(утверждение, помогающее создать положительный психологический настрой. '
                           'Многократное повторение воздействует на подсознание и помогает создать новую модель мышления) - ', grey, 'I'],
                            [DB_get_value(DB['Задача для ТС'], 'H', num_data['goal_ts']), blue, 'I']], end_padding=True)
    ##
    pdf_print_block(pdf, f'Любовь - {num_data["love"]}', '(характер вазимоотношений с партнером)', DB_get_value(DB['Любовь'], 'B', num_data['love']))
    ##
    pdf_print_block(pdf, f'Деньги/доходы - {num_data["m_in"]}', '(источник дохода)', DB_get_value(DB['Деньги доходы'], 'B', num_data['m_in']))
    ##
    pdf_print_block(pdf, f'Забота - {num_data["care"]}', '(как обычно проявляете свою заботу и как, в Вашем случае, необходимо заботиться о людях)',
                    DB_get_value(DB['Забота'], 'B', num_data['care']))
    ##
    pdf_print_block(pdf, f'Возможности - {num_data["poss"]}')
    with pdf.table(line_height=1.59 * pdf.font_size, headings_style=headings_style, num_heading_rows=2,
                   borders_layout='NO_HORIZONTAL_LINES', text_align='L') as table:
        row = table.row()
        row.cell('Врожденные действия (форма поведения)', align='C', padding=3)
        row.cell('Измененные действия (форма поведения)', align='C', padding=3)

        row = table.row()
        pdf.set_font(style='I')
        row.cell('(неосознаваемые поведенческие паттерны. Готовность действовать определённым образом)', align='C', padding=(0, 3, 3, 3))
        row.cell('(сознательное изменение действий для максимальной реализации)', align='C', padding=(0, 3, 3, 3))

        row = table.row()
        pdf.set_font(style='BI')
        pdf.set_text_color(blue)
        row.cell('Деструктивные действия (форма поведения):', padding=(3, 3, 0, 3))
        row.cell('Конструктивные действия (форма поведения):', padding=(3, 3, 0, 3))

        row = table.row()
        pdf.set_font(style='')
        pdf.set_text_color(black)
        row.cell(DB_get_value(DB['Возможности'], 'B', num_data['poss']), padding=(0, 3, 3, 3))
        row.cell(DB_get_value(DB['Возможности'], 'D', num_data['poss']), padding=(0, 3, 3, 3), rowspan=3, v_align='T')

        row = table.row()
        pdf.set_font(style='BI')
        pdf.set_text_color(blue)
        row.cell('Конструктивные действия (форма поведения):', padding=(3, 3, 0, 3))

        row = table.row()
        pdf.set_font(style='')
        pdf.set_text_color(black)
        row.cell(DB_get_value(DB['Возможности'], 'C', num_data['poss']), padding=(0, 3, 3, 3))
    ##
    pdf_print_block(pdf, f'Мировоззрение - {num_data["wv"]}',
                    '(совокупная система взглядов, принципов и ценностей, определяющая мотивы поведения)',
                    DB_get_value(DB['Мировоззрение'], 'B', num_data['wv']))
    ##
    vector_subtitle='(показатель направленности в жизни, т.е. совокупность энергий, через которые человек приходит либо к стагнации и разрушению, либо к реализации в жизни. Самореализация — процесс, который заключается в реализации человеком своих способностей, потенциалов и талантов, в каком-либо виде деятельности)'
    pdf_print_block(pdf, f'Вектор жизни - {num_data["vector"]}', vector_subtitle)
    pdf_print_colored_text(pdf, [[DB_get_value(DB['Вектор жизни'], 'B', num_data['vector']), red, 'BI']], end_padding=True)
    pdf_print_colored_text(pdf, [['Стагнация (отсутсвие развития): ', blue, 'BI'],
                                  [DB_get_value(DB['Вектор жизни'], 'C', num_data['vector']), black, '']], end_padding=True)
    pdf_print_colored_text(pdf, [['Самореализация (через что происходит развитие потенциала личности): ', blue, 'BI'],
                                  [DB_get_value(DB['Вектор жизни'], 'D', num_data['vector']), black, '']], end_padding=True)
    ##
    pdf_print_block(pdf, f'Ключевая цель/долг - {num_data["key_goal"]}',
                    '(направление, которое даёт стабильность в жизни)',
                    DB_get_value(DB['Ключевая цель'], 'B', num_data['key_goal']))
    ##
    pdf_print_block(pdf, f'Первоисточник проблем - {num_data["pp"]}')
    #
    # Переделать под новый текст от т Гули
    #
    with pdf.table(line_height=1.59 * pdf.font_size, headings_style=headings_style, v_align='T', padding=3) as table:
        row = table.row()
        row.cell('Что создает проблемы/трудности', align='C')
        row.cell('Над чем необходимо работать', align='C')

        row = table.row()
        row.cell(DB_get_value(DB['Первоисточник проблем'], 'B', num_data['pp']), align='L')
        row.cell(DB_get_value(DB['Первоисточник проблем'], 'C', num_data['pp']), align='L')
    ##
    pdf_print_block(pdf, 'Матрица качеств и компетенций', '(врожденные качетва и компетенции личности)')
    pdf.set_font(style='B')

    comp_matrix, matrix_lines = calc_competence_matrix(input_date)

    with pdf.table(comp_matrix, first_row_as_headings=False, padding = 5, line_height=1.59 * pdf.font_size, text_align='C', width=60):
        pass
    pdf.set_font(style='')
    ##
    pdf_print_block(pdf, 'Линии в матрице', '(сочетание качеств и компетенций формируют характерные черты личности и врожденный потенциал личности)')
    for ln in matrix_lines.values():
        DB_row = int(ln[0]) - 2  # -2 нужно, тк получение данных с этого листа отличается от остальных

        pdf_print_colored_text(pdf, [[DB_get_value(DB['Матрица качеств и компетенций'], 'B', DB_row), blue, 'BI'],
                                      [' (где ' + DB_get_value(DB['Матрица качеств и компетенций'], 'C', DB_row), black, ''],
                                      ['. ' + DB_get_value(DB['Матрица качеств и компетенций'], 'D', DB_row) + ')', black, '']], True)
    ##
    pdf_print_block(pdf, 'В Вашей матрице заложены следующие качества/компетенции:')
    num = 0
    for x in range(3):
        for y in range(2, -1, -1):
            num += 1
            if comp_matrix[y][x] != ' ':
                pdf_print_text(pdf, '•  ' + DB_get_value(DB['Цифры в матрице по отдельности'], 'B', num),
                               padding=False)
    ##
    pdf_print_block(pdf, 'В Вашей матрице отсутствуют следующие качества/компетенции:')
    num = 0
    for x in range(3):
        for y in range(2, -1, -1):
            num += 1
            if comp_matrix[y][x] == ' ':
                pdf_print_text(pdf, '•  ' + DB_get_value(DB['Цифры в матрице по отдельности'], 'B', num),
                               padding=False)
    ##
    pdf_print_block(pdf, 'Рекомендации по наработке отсутствующих в матрице качеств и компетенций:')
    pdf_print_colored_text(pdf, [['Общие рекомендации:', blue, 'BI']])
    pdf.set_xy(x=pdf.l_margin + 3, y=pdf.y + 12)
    pdf_print_colored_text(pdf, [['•  Стремитесь к выполнению Задачи для трансформации сознания!', red, '']])
    pdf.set_xy(x=pdf.l_margin + 3, y=pdf.y + 6)
    pdf_print_text(pdf, '•  В течение дня пейте теплую воду, до 45 градусов С (на 30 кг – 1 литр воды)', padding=False)
    pdf.set_xy(x=pdf.l_margin + 3, y=pdf.y)
    pdf_print_text(pdf, '•  Ежедневно ходите от 6 км. и больше со скоростью 5 км/час и выше', padding=False)
    pdf.set_xy(x=pdf.l_margin + 3, y=pdf.y + 12)
    ##
    num = 0
    for x in range(3):
        for y in range(2, -1, -1):
            num += 1
            if comp_matrix[y][x] == ' ':
                pdf_print_colored_text(pdf, [['Как наработать качества/компетенции «' + str(num) + '»:', blue, 'BI']])
                pdf.set_xy(x=pdf.l_margin + 3, y=pdf.y + 12)
                pdf_print_text(pdf, make_list(DB_get_value(DB['Цифры в матрице по отдельности'], 'C', num)), padding=False, align='L')
                pdf.set_xy(x=pdf.l_margin + 3, y=pdf.y + 12)
    del num
    ##
    current_year = date.today().year
    personal_year = spec_sum(str(num_data['day']) + str(num_data['month']) + str(current_year))
    pdf_print_block(pdf, f'Личный год {current_year}г. - {personal_year}')
    pdf_print_colored_text(pdf, [[DB_get_value(DB[f'Личный год {current_year}'], 'B', personal_year), red, 'BI']], True)
    pdf_print_colon_text(pdf, 'В «плюсе»: ', DB_get_value(DB[f'Личный год {current_year}'], 'C', personal_year))
    pdf_print_colon_text(pdf, 'В «минусе»: ', DB_get_value(DB[f'Личный год {current_year}'], 'D', personal_year))
    pdf_print_colored_text(pdf, [['Рекомендации на этот год: ', red, 'BI'], [DB_get_value(DB[f'Личный год {current_year}'], 'E', personal_year), grey, 'I']], True)
    ##
    pdf.set_text_color(red)
    pdf.set_font(style='B')
    pdf_print_text(pdf, f'{name}, я от всей души желаю Вам успеха!', align='C')
    pdf_print_text(pdf, 'С Уважением, Гульнара', align='C')

    pdf.output(name + '.' + input_date + '.pdf')
    os.system(name + '.' + input_date + '.pdf')



class DigitalPsychologyCalcApp(App):
    Window.size = (228, 512)
    Window.clearcolor = (.11, .3, .25, 1)
    Config.set('graphics', 'resizable', '0')
    def build(self):
        self.DB = load_DB()
        wrapper = BoxLayout(orientation='vertical', padding=[20, 100, 20, 20])
        main_block = BoxLayout(orientation='vertical', spacing=10)
        self.name_input = TextInput(text='Daniil', hint_text='Имя', multiline=False, size_hint=[1, None], height=35)
        self.date_input = TextInput(text='27.06.2005', hint_text='дд.мм.гггг', multiline=False, size_hint=(1, None), height=35)

        main_block.add_widget(Label(text='Имя', size_hint=[1, None], height=20, halign='left', text_size=[188, 20]))
        main_block.add_widget(self.name_input)
        main_block.add_widget(Label(text='Дата рождения', size_hint=[1, None], height=20, halign='left', text_size=[188, 20]))
        main_block.add_widget(self.date_input)
        main_block.add_widget(Button(text='Рассчитать', on_press=self.calc_btn, size_hint=(1, None), height=35))

        bottom_block = AnchorLayout(anchor_x='center', anchor_y='bottom')
        bottom_block.add_widget(Button(text='Добавить данные', on_press=self.add_data_btn, size_hint=(1, None), height=35))

        wrapper.add_widget(main_block)
        wrapper.add_widget(bottom_block)

        return wrapper

    # Нажатие кнопки "рассчитать"
    def calc_btn(self, instance):
        name = self.name_input.text
        date = self.date_input.text

        self.name_input.hint_text = 'имя'
        self.name_input.hint_text_color = [.5, .5, .5, 1]
        self.date_input.hint_text = 'дд.мм.гггг'
        self.date_input.hint_text_color = [.5, .5, .5, 1]

        if name == '':
            self.name_input.hint_text = 'Введите имя'
            self.name_input.hint_text_color = [1, 0, 0, 1]
            self.name_input.text = ''
        elif not check_date(date):
            self.date_input.hint_text = 'Введите дату в соответствии с макетом: дд.мм.гггг'
            self.date_input.hint_text_color = [1, 0, 0, 1]
            self.date_input.text = ''
        else:
            create_pdf(date, name, self.DB)

    # кнопка для добавления данных в excel
    def add_data_btn(self, instance):
        os.system('PsyCalcData.xlsx')


if __name__ == "__main__":
    DigitalPsychologyCalcApp().run()
